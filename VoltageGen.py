import time
import os
from openpyxl import Workbook, load_workbook
from adafruit_ina219 import INA219,ADCResolution,BusVoltageRange
import busio
import board
from datetime import datetime

#initialize the i2c bus and addresses for the ina219 sensors
i2c= busio.I2C(board.SCL, board.SDA)
genSensor = INA219(i2c, 0x41)

#sets the adc resolution and voltage range for more accurate data collection
genSensor.bus_adc_resoltion = ADCResolution.ADCRES_12BIT_32S
genSensor.shunt_adc_resoltion = ADCResolution.ADCRES_12BIT_32S
genSensor.bus_voltage_range = BusVoltageRange.RANGE_16V

#sets the location to be stored to be on a USB
#filepath = "/media/egr385/KEY2SUCCESS/RC_Sensor_Data.xlsx"
filepath = "/media/egr385/ESD-USB/Sensors_Data.xlsx"    

try:
    if os.path.exists(filepath):    
        #look to see if there is a workbook already created, if yes then load the workbook
        wb = load_workbook(filepath)
        print("loaded workbook")
    else:   
        raise FileNotFoundError
except(FileNotFoundError, KeyError):
    #If File not found create a new sheet.
    wb = Workbook()
    wb.save(filepath)
    print("created new wb")

#creates a unique sheet name for each trial
sheet_name = datetime.now().strftime("Track_Data_%d_%H%M%S") 
ws = wb.create_sheet(title=sheet_name)

#set the headers in the excel sheet
ws.append(["Timestamp","genVoltage (V)", "genCurrent (A)", "genPower (W)"])
wb.save(filepath)

try:
    for timestamp in range(600):    #roughly 10 min
        #get readings from the fan generator, and calibration
        genVoltage = abs(genSensor.bus_voltage - 0.868)*10
        genCurrent = abs(genSensor.current/1000)
        genPower = genVoltage*genCurrent
            
        print("Timestamp: ",timestamp)
        print("Gen")
        print("Voltage: {:6.3f}V".format(genVoltage), "Current: {:7.4f}A".format(genCurrent), "Power: {:6.3f}W".format(genPower))

        #save data between every literation to reduce the risk of the data being lost
        ws.append([genVoltage,genCurrent,genPower])
        wb.save(filepath) 
        time.sleep(0.5)
        
except KeyboardInterrupt:
    """This will only occure when False"""
    #key board interrupt is crtl + c 
    print("Data logging stoped")
    #saves the data at the very end to ensure all data is saved

finally:
    wb.save(filepath)
    wb.close() 
    print("Data saved to sensor_data.xlsx")
